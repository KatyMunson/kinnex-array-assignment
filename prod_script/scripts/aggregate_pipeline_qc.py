#!/usr/bin/env python3
"""
aggregate_pipeline_qc.py

Aggregates QC metrics across all pipeline stages (Skera, Lima pass 1,
assignment, split, Lima pass 2, refine) into per-stage TSV files and a
formatted Excel workbook.

Handles both single-library and multi-library samples in the same run.
Single-library samples have no Lima pass 1, assignment, or split data;
those columns appear as blank/None for those rows. All samples have Skera,
Lima pass 2, refine, and summary data.

A processing_mode column (SINGLE-LIBRARY / MULTI-LIBRARY) is added to the
Refine tab and propagated to the Summary tab to make the per-sample
processing decision visible in the workbook.

Usage: python aggregate_pipeline_qc.py <results_dir> <manifest> <out_dir>

Author:  KM
Created: 2026-02
"""

import sys
import os
import re
import json
import glob
import argparse
import pandas as pd
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# utils.py lives in the same scripts/ directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import utils

parser = argparse.ArgumentParser()
parser.add_argument("results_dir")
parser.add_argument("manifest")
parser.add_argument("out_dir")
parser.add_argument("--posteriors-dir", default=None,
                    help="Directory containing per-sample posterior PNG files "
                         "(results/qc/posteriors/). If provided, PNGs are "
                         "embedded in the Posteriors sheet of the Excel workbook.")
args = parser.parse_args()

results_dir    = Path(args.results_dir)
manifest       = Path(args.manifest)
out_dir        = Path(args.out_dir)
posteriors_dir = Path(args.posteriors_dir) if args.posteriors_dir else None
out_dir.mkdir(parents=True, exist_ok=True)

# ── Load manifest and resolve modes ───────────────────────────────────────────
# parse_manifest performs mode resolution and logs decisions, matching what
# the Snakefile did at submission time. This ensures the QC report reflects
# the same mode decisions as the pipeline run.
_mf = utils.parse_manifest(manifest)

samples      = _mf["samples"]
arrays_files = _mf["array_files"]
modes        = _mf["modes"]        # { sample: "single" | "multi" }
libraries    = _mf["libraries"]    # { sample: [lib_name, ...] }

def mode_label(sample):
    """Return human-readable processing mode for Excel display."""
    return "SINGLE-LIBRARY" if modes.get(sample) == utils.SINGLE else "MULTI-LIBRARY"

print(f"Samples: {samples}")


# ── Helper functions ───────────────────────────────────────────────────────────

def load_arrays(arrays_path):
    """Return dict of library -> set of expected barcodes from an arrays file."""
    expected = {}
    try:
        with open(arrays_path) as f:
            for line in f:
                parts = line.strip().split("\t")
                if len(parts) < 3:
                    continue
                expected[parts[0]] = set(parts[2:])
    except Exception as e:
        print(f"WARNING: Could not parse arrays file {arrays_path}: {e}")
    return expected


def parse_lima_counts(counts_path):
    """Return dict of IdxFirstNamed -> (Counts, MeanScore) from a lima.counts file."""
    result = {}
    try:
        df = pd.read_csv(counts_path, sep="\t")
        for _, row in df.iterrows():
            result[row["IdxFirstNamed"]] = (int(row["Counts"]), float(row["MeanScore"]))
    except Exception as e:
        print(f"WARNING: Could not parse counts file {counts_path}: {e}")
    return result


def parse_skera_summary(summary_path):
    """Parse a .skera.summary.csv file into a dict of key -> value strings."""
    metrics = {}
    try:
        with open(summary_path) as f:
            for line in f:
                if "," in line:
                    k, v = line.strip().split(",", 1)
                    metrics[k.strip()] = v.strip()
    except Exception:
        pass
    return metrics


def parse_lima_summary(summary_path):
    """Parse a .lima.summary file into a dict of key -> value strings."""
    metrics = {}
    try:
        with open(summary_path) as f:
            for line in f:
                line = line.strip()
                if ":" in line:
                    k, v = line.split(":", 1)
                    metrics[k.strip()] = v.strip()
    except Exception:
        pass
    return metrics


def get_row(df, sample):
    """Return first matching row for a sample, or None if not found."""
    if df is None or len(df) == 0:
        return None
    rows = df.loc[df["sample"] == sample]
    return rows.iloc[0] if len(rows) else None


def get_rows(df, sample):
    """Return all rows for a sample as a DataFrame."""
    if df is None or len(df) == 0:
        return pd.DataFrame()
    return df.loc[df["sample"] == sample]


# ══════════════════════════════════════════════════════════════════════════════
# Stage 1: Skera — all samples
# ══════════════════════════════════════════════════════════════════════════════
print("Parsing skera...")
skera_rows = []
for s in samples:
    summary_path = results_dir / "skera" / f"{s}.skera.summary.csv"
    m = parse_skera_summary(summary_path)
    skera_rows.append({
        "sample":            s,
        "processing_mode":   mode_label(s),
        "input_reads":       int(m.get("Input Reads", 0)),
        "segmented_reads":   int(m.get("Segmented Reads (S-Reads)", 0)),
        "mean_sread_length": float(m.get("Mean Length of S-Reads", 0)),
        "pct_full_array":    float(m.get("Percentage of Reads with Full Array", 0)),
        "mean_array_size":   float(m.get("Mean Array Size (Concatenation Factor)", 0)),
    })

df_skera = pd.DataFrame(skera_rows)
df_skera.to_csv(out_dir / "qc_skera.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Stage 2: Lima pass 1 — multi-library samples only
# Single-library samples have no Lima pass 1 and are omitted from this tab.
# ══════════════════════════════════════════════════════════════════════════════
print("Parsing lima pass 1 (multi-library samples only)...")
lima1_rows    = []
lima1_bc_rows = []

for s in [s for s in samples if modes[s] == utils.MULTI]:
    lima_dir      = results_dir / "lima" / s
    summary_files = glob.glob(str(lima_dir / "*.lima.summary"))
    counts_files  = glob.glob(str(lima_dir / "*.lima.counts"))
    if not summary_files:
        continue
    m      = parse_lima_summary(summary_files[0])
    counts = parse_lima_counts(counts_files[0]) if counts_files else {}

    reads_in   = int(re.sub(r"[^\d]", "", m.get("Reads input", "0")) or 0)
    reads_pass = int(re.sub(r"[^\d]", "", m.get("Reads above all thresholds (A)", "0")) or 0)
    reads_fail = reads_in - reads_pass
    detected   = {bc for bc in counts if bc != "IsoSeqX_3p"}

    lima1_rows.append({
        "sample":            s,
        "reads_in":          reads_in,
        "reads_passed":      reads_pass,
        "reads_failed":      reads_fail,
        "pct_passed":        round(100 * reads_pass / reads_in, 2) if reads_in else None,
        "barcodes_detected": len(detected),
        "barcodes":          ",".join(sorted(detected)),
    })

    for bc, (cnt, score) in sorted(counts.items()):
        if bc == "IsoSeqX_3p":
            continue
        lima1_bc_rows.append({
            "sample":     s,
            "barcode":    bc,
            "reads":      cnt,
            "mean_score": score,
        })

df_lima1    = pd.DataFrame(lima1_rows)    if lima1_rows    else pd.DataFrame()
df_lima1_bc = pd.DataFrame(lima1_bc_rows) if lima1_bc_rows else pd.DataFrame()
df_lima1.to_csv(out_dir / "qc_lima1.tsv", sep="\t", index=False)
df_lima1_bc.to_csv(out_dir / "qc_lima1_barcodes.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Stage 3: Array assignment — multi-library samples only
# ══════════════════════════════════════════════════════════════════════════════
print("Parsing array assignment (multi-library samples only)...")
assign_rows     = []
assign_lib_rows = []

for s in [s for s in samples if modes[s] == utils.MULTI]:
    assign_file = results_dir / "assigned" / f"{s}.txt"
    if not assign_file.exists():
        continue
    df = utils.load_assignments_df(assign_file)
    cls_counts = df["Classification"].value_counts().to_dict()
    total_zmws = len(df)

    assign_rows.append({
        "sample":        s,
        "total_zmws":    total_zmws,
        "HIGH_CONF":     cls_counts.get("HIGH_CONF", 0),
        "LOW_CONF":      cls_counts.get("LOW_CONF", 0),
        "UNASSIGNED":    cls_counts.get("UNASSIGNED", 0),
        "pct_high_conf": round(100 * cls_counts.get("HIGH_CONF", 0) / total_zmws, 2)
                         if total_zmws else None,
    })

    hc = df[df["Classification"] == "HIGH_CONF"]
    for lib, grp in hc.groupby("Assigned_Array"):
        assign_lib_rows.append({
            "sample":    s,
            "library":   lib,
            "zmw_count": len(grp),
        })

df_assign     = pd.DataFrame(assign_rows)     if assign_rows     else pd.DataFrame()
df_assign_lib = pd.DataFrame(assign_lib_rows) if assign_lib_rows else pd.DataFrame()
df_assign.to_csv(out_dir / "qc_assign.tsv", sep="\t", index=False)
df_assign_lib.to_csv(out_dir / "qc_assign_by_library.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Stage 4: Split Skera — all samples
# Single-library samples have a minimal split_skera_qc.tsv written by the
# passthrough rule; it contains the library.HIGH_CONF read count only.
# ══════════════════════════════════════════════════════════════════════════════
print("Parsing split skera...")
split_rows = []
for s in samples:
    qc_file = results_dir / "split_skera" / s / "split_skera_qc.tsv"
    if not qc_file.exists():
        continue
    df = pd.read_csv(qc_file, sep="\t")
    for _, row in df.iterrows():
        key = row["library"]
        if key == "unassigned":
            library, confidence = "unassigned", "unassigned"
        elif "." in key:
            library, confidence = key.rsplit(".", 1)
            confidence = confidence.lower().replace("_conf", "conf")
        else:
            library, confidence = key, "unknown"
        split_rows.append({
            "sample":          s,
            "processing_mode": mode_label(s),
            "library":         library,
            "confidence":      confidence,
            "read_count":      row["read_count"],
        })

df_split = pd.DataFrame(split_rows) if split_rows else pd.DataFrame()
df_split.to_csv(out_dir / "qc_split_skera.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Stage 5: Lima pass 2 — all samples
# ══════════════════════════════════════════════════════════════════════════════
print("Parsing lima pass 2...")
lima2_rows    = []
lima2_bc_rows = []

for s in samples:
    lima2_base = results_dir / "lima2" / s
    if not lima2_base.exists():
        continue
    expected_by_lib = load_arrays(arrays_files[s])
    for lib_dir in sorted(lima2_base.iterdir()):
        if not lib_dir.is_dir():
            continue
        lib      = lib_dir.name
        expected = expected_by_lib.get(lib, set())
        for conf_dir in sorted(lib_dir.iterdir()):
            if not conf_dir.is_dir():
                continue
            conf          = conf_dir.name
            summary_files = glob.glob(str(conf_dir / "*.lima.summary"))
            counts_files  = glob.glob(str(conf_dir / "*.lima.counts"))
            if not summary_files:
                continue
            m      = parse_lima_summary(summary_files[0])
            counts = parse_lima_counts(counts_files[0]) if counts_files else {}

            reads_in   = int(re.sub(r"[^\d]", "", m.get("Reads input", "0")) or 0)
            reads_pass = int(re.sub(r"[^\d]", "", m.get("Reads above all thresholds (A)", "0")) or 0)
            reads_fail = reads_in - reads_pass
            detected   = {bc for bc in counts if bc != "IsoSeqX_3p"}
            unexpected = detected - expected
            missing    = expected - detected

            lima2_rows.append({
                "sample":              s,
                "processing_mode":     mode_label(s),
                "library":             lib,
                "confidence":          conf,
                "reads_in":            reads_in,
                "reads_passed":        reads_pass,
                "reads_failed":        reads_fail,
                "pct_passed":          round(100 * reads_pass / reads_in, 2) if reads_in else None,
                "barcodes_expected":   len(expected),
                "barcodes_detected":   len(detected),
                "barcodes_unexpected": len(unexpected),
                "barcodes_missing":    len(missing),
                "unexpected_list":     ",".join(sorted(unexpected)) if unexpected else "",
                "missing_list":        ",".join(sorted(missing))    if missing    else "",
            })

            for bc, (cnt, score) in sorted(counts.items()):
                if bc == "IsoSeqX_3p":
                    continue
                lima2_bc_rows.append({
                    "sample":          s,
                    "processing_mode": mode_label(s),
                    "library":         lib,
                    "confidence":      conf,
                    "barcode":         bc,
                    "reads":           cnt,
                    "mean_score":      score,
                    "expected":        bc in expected,
                })

df_lima2    = pd.DataFrame(lima2_rows)    if lima2_rows    else pd.DataFrame()
df_lima2_bc = pd.DataFrame(lima2_bc_rows) if lima2_bc_rows else pd.DataFrame()
df_lima2.to_csv(out_dir / "qc_lima2.tsv", sep="\t", index=False)
df_lima2_bc.to_csv(out_dir / "qc_lima2_barcodes.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Stage 6: Refine (FLNC) — all samples
# processing_mode is the primary column where the single/multi decision is
# surfaced for the user; it appears here and in Summary.
# ══════════════════════════════════════════════════════════════════════════════
print("Parsing refine...")
BARCODES = [f"bc{i:02d}" for i in range(1, 13)]

refine_counts = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))

json_files = glob.glob(
    str(results_dir / "refine" / "*" / "*" / "*" / "*.flnc.filter_summary.report.json")
)
for jf in sorted(json_files):
    p          = Path(jf)
    sample_    = p.parts[-4]
    library    = p.parts[-3]
    confidence = p.parts[-2]
    match      = re.search(r"_(bc\d{2})_", p.name)
    if not match:
        continue
    bc = match.group(1)
    with open(p) as f:
        data = json.load(f)
    for attr in data.get("attributes", []):
        if attr.get("id") == "num_reads_flnc_polya":
            refine_counts[sample_][library][confidence][bc] += attr["value"]

refine_rows = []
for sample_ in sorted(refine_counts):
    for library in sorted(refine_counts[sample_]):
        for confidence in sorted(refine_counts[sample_][library]):
            row = {
                "sample":          sample_,
                "processing_mode": mode_label(sample_),
                "library":         library,
                "confidence":      confidence,
            }
            for bc in BARCODES:
                row[bc] = refine_counts[sample_][library][confidence].get(bc, 0)
            row["total_flnc"] = sum(refine_counts[sample_][library][confidence].values())
            refine_rows.append(row)

df_refine = pd.DataFrame(refine_rows) if refine_rows else pd.DataFrame()
df_refine.to_csv(out_dir / "qc_refine.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Master summary: read tracking across all stages
# ══════════════════════════════════════════════════════════════════════════════
print("Building master summary...")
summary_rows = []

for s in samples:
    sk     = get_row(df_skera, s)
    sk_in  = sk["input_reads"]     if sk is not None else None
    sk_out = sk["segmented_reads"] if sk is not None else None

    l1 = get_row(df_lima1, s) if len(df_lima1) > 0 else None
    a  = get_row(df_assign, s) if len(df_assign) > 0 else None

    sp               = get_rows(df_split, s)
    sp_highconf      = sp.loc[(sp["library"] != "unassigned") & (sp["confidence"] == "highconf")] if len(sp) else pd.DataFrame()
    sp_lowconf       = sp.loc[(sp["library"] != "unassigned") & (sp["confidence"] == "lowconf")]  if len(sp) else pd.DataFrame()
    sp_unassigned    = sp.loc[sp["library"] == "unassigned"] if len(sp) else pd.DataFrame()
    split_highconf   = int(sp_highconf["read_count"].sum())   if len(sp_highconf)   else None
    split_lowconf    = int(sp_lowconf["read_count"].sum())    if len(sp_lowconf)    else None
    split_unassigned = int(sp_unassigned["read_count"].sum()) if len(sp_unassigned) else None
    split_total      = (split_highconf or 0) + (split_lowconf or 0) + (split_unassigned or 0)
    split_pct_assigned = (
        round(100 * ((split_highconf or 0) + (split_lowconf or 0)) / split_total, 2)
        if split_total else None
    )

    l2_hc        = df_lima2.loc[(df_lima2["sample"] == s) & (df_lima2["confidence"] == "highconf")] if len(df_lima2) else pd.DataFrame()
    l2_lc        = df_lima2.loc[(df_lima2["sample"] == s) & (df_lima2["confidence"] == "lowconf")]  if len(df_lima2) else pd.DataFrame()
    lima2_hc_in  = int(l2_hc["reads_in"].sum())      if len(l2_hc) else None
    lima2_hc_pass = int(l2_hc["reads_passed"].sum())  if len(l2_hc) else None
    lima2_lc_in  = int(l2_lc["reads_in"].sum())      if len(l2_lc) else None
    lima2_lc_pass = int(l2_lc["reads_passed"].sum())  if len(l2_lc) else None

    rf_hc     = df_refine.loc[(df_refine["sample"] == s) & (df_refine["confidence"] == "highconf")] if len(df_refine) else pd.DataFrame()
    rf_lc     = df_refine.loc[(df_refine["sample"] == s) & (df_refine["confidence"] == "lowconf")]  if len(df_refine) else pd.DataFrame()
    flnc_hc   = int(rf_hc["total_flnc"].sum()) if len(rf_hc) else None
    flnc_lc   = int(rf_lc["total_flnc"].sum()) if len(rf_lc) else None
    flnc_total = (flnc_hc or 0) + (flnc_lc or 0) or None

    summary_rows.append({
        "sample":                s,
        "processing_mode":       mode_label(s),
        "skera_input_reads":     sk_in,
        "skera_segmented_reads": sk_out,
        "skera_pct_full_array":  sk["pct_full_array"]  if sk is not None else None,
        "skera_mean_array_size": sk["mean_array_size"] if sk is not None else None,
        # Lima pass 1: blank for single-library samples
        "lima1_reads_in":        l1["reads_in"]     if l1 is not None else None,
        "lima1_reads_passed":    l1["reads_passed"] if l1 is not None else None,
        "lima1_pct_passed":      l1["pct_passed"]   if l1 is not None else None,
        # Assignment: blank for single-library samples
        "assign_high_conf":      a["HIGH_CONF"]     if a is not None else None,
        "assign_low_conf":       a["LOW_CONF"]      if a is not None else None,
        "assign_unassigned":     a["UNASSIGNED"]    if a is not None else None,
        "assign_pct_high_conf":  a["pct_high_conf"] if a is not None else None,
        # Split
        "split_highconf_reads":    split_highconf,
        "split_lowconf_reads":     split_lowconf,
        "split_unassigned_reads":  split_unassigned,
        "split_pct_assigned":      split_pct_assigned,
        # Lima pass 2
        "lima2_hc_reads_in":    lima2_hc_in,
        "lima2_hc_reads_passed": lima2_hc_pass,
        "lima2_hc_pct_passed":  round(100 * lima2_hc_pass / lima2_hc_in, 2) if lima2_hc_in else None,
        "lima2_lc_reads_in":    lima2_lc_in,
        "lima2_lc_reads_passed": lima2_lc_pass,
        "lima2_lc_pct_passed":  round(100 * lima2_lc_pass / lima2_lc_in, 2) if lima2_lc_in else None,
        # Refine
        "flnc_highconf_reads":  flnc_hc,
        "flnc_lowconf_reads":   flnc_lc,
        "flnc_total_reads":     flnc_total,
        "overall_yield_pct":    round(100 * flnc_total / sk_out, 2) if (sk_out and flnc_total) else None,
        "highconf_yield_pct":   round(100 * flnc_hc / sk_out, 2)    if (sk_out and flnc_hc)    else None,
    })

df_summary = pd.DataFrame(summary_rows)
df_summary.to_csv(out_dir / "qc_summary.tsv", sep="\t", index=False)


# ══════════════════════════════════════════════════════════════════════════════
# Excel workbook
# ══════════════════════════════════════════════════════════════════════════════
print("Writing Excel workbook...")

HEADER_FILL  = PatternFill("solid", start_color="366092", end_color="366092")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
NORMAL_FONT  = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    bottom=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin",  color="B8CCE4"),
)
# Mode column highlight colours
SINGLE_FILL  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")  # light green
MULTI_FILL   = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")  # light blue


def write_df_to_sheet(ws, df, title=None, mode_col=None):
    """
    Write a DataFrame to a worksheet with standard formatting.

    If mode_col names a column present in df, cells in that column are
    highlighted: SINGLE-LIBRARY in green, MULTI-LIBRARY in blue. This makes
    the processing mode immediately visible when scanning the sheet.
    """
    if df is None or len(df) == 0:
        if title:
            ws.cell(row=1, column=1, value=title).font = Font(
                name="Arial", size=12, bold=True, color="17375E"
            )
        ws.cell(row=3, column=1, value="(no data for this stage in this run)")
        return

    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(
            name="Arial", size=12, bold=True, color="17375E"
        )
        start_row = 3

    col_names = list(df.columns)

    for col_idx, col_name in enumerate(col_names, start=1):
        cell           = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        alt = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (col_name, val) in enumerate(zip(col_names, row), start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = NORMAL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border    = THIN_BORDER
            if mode_col and col_name == mode_col:
                cell.fill = SINGLE_FILL if str(val) == "SINGLE-LIBRARY" else MULTI_FILL
            elif alt:
                cell.fill = alt

    for col_idx, col_name in enumerate(col_names, start=1):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# Sheet definitions: (tab_name, dataframe, title, mode_col_or_None)
sheets = [
    ("Summary",        df_summary,     "Pipeline QC — Master Read Tracking",           "processing_mode"),
    ("Skera",          df_skera,       "Stage 1: Skera Segmentation",                  "processing_mode"),
    ("Lima_Pass1",     df_lima1,       "Stage 2: Lima Pass 1 — Multi-Library Only",    None),
    ("Lima1_Barcodes", df_lima1_bc,    "Stage 2: Lima Pass 1 — Per Barcode Counts",    None),
    ("Assign",         df_assign,      "Stage 3: Array Assignment — Per Sample",       None),
    ("Assign_Library", df_assign_lib,  "Stage 3: Array Assignment — Per Library",      None),
    ("Split_Skera",    df_split,       "Stage 4: Split Skera by Library",              "processing_mode"),
    ("Lima_Pass2",     df_lima2,       "Stage 5: Lima Pass 2 — Barcode Check",         "processing_mode"),
    ("Lima2_Barcodes", df_lima2_bc,    "Stage 5: Lima Pass 2 — Per Barcode Counts",    "processing_mode"),
    ("Refine",         df_refine,      "Stage 6: IsoSeq Refine — FLNC PolyA Reads",   "processing_mode"),
]

wb = Workbook()
for i, (sheet_name, df, title, mode_col) in enumerate(sheets):
    ws = wb.active if i == 0 else wb.create_sheet(sheet_name)
    if i == 0:
        ws.title = sheet_name
    write_df_to_sheet(ws, df, title=title, mode_col=mode_col)

# ── Posteriors sheet ───────────────────────────────────────────────────────────
# Embed one PNG per multi-library sample, stacked vertically.
# Each image is preceded by the sample name as a label.
# Skipped gracefully if no posteriors directory was provided or no PNGs exist.
if posteriors_dir and posteriors_dir.exists():
    png_files = sorted(posteriors_dir.glob("*_posteriors.png"))
    if png_files:
        ws_post = wb.create_sheet("Posteriors")
        ws_post.cell(row=1, column=1, value="Posterior Distributions — Multi-Library Samples").font = Font(
            name="Arial", size=12, bold=True, color="17375E"
        )
        # Each PNG is placed at a fixed height of ~300px = ~225pt = ~30 rows.
        # Images are anchored by cell reference; row spacing keeps them from
        # overlapping. Adjust IMG_ROW_HEIGHT if your PNGs are a different size.
        IMG_ROW_HEIGHT = 32   # rows to advance between images
        current_row    = 3
        for png_path in png_files:
            sample_name = png_path.stem.replace("_posteriors", "")
            ws_post.cell(row=current_row, column=1, value=sample_name).font = Font(
                name="Arial", size=10, bold=True, color="366092"
            )
            img = XLImage(str(png_path))
            # Scale to a consistent display width of ~900px while preserving
            # aspect ratio. The PNG is 150dpi at 10in wide = 1500px; scale to
            # roughly half for comfortable workbook display.
            img.width  = 750
            img.height = int(img.width * 5 / 10)  # 10x5in original aspect ratio
            ws_post.add_image(img, f"A{current_row + 1}")
            current_row += IMG_ROW_HEIGHT
        print(f"[pipeline_qc] Embedded {len(png_files)} posterior plot(s) in Posteriors sheet.")
    else:
        print(f"[pipeline_qc] WARNING: posteriors dir {posteriors_dir} exists but contains no PNGs.")
else:
    if posteriors_dir:
        print(f"[pipeline_qc] WARNING: posteriors dir {posteriors_dir} not found — skipping Posteriors sheet.")

wb.save(out_dir / "pipeline_qc.xlsx")
print(f"Done. Outputs written to {out_dir}")
